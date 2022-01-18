from matplotlib import pyplot # Импортировать модуль pyplot из библиотеки matplotlib
from openpyxl import load_workbook # Импортировать модуль load_workbook из библиотеки openpyxl
wb = load_workbook('data_analysis_lab.xlsx') # Загрузить таблицу Excel из файла "data_analysis_lab.xlsx" в переменную wb
sheet = wb['Data'] # Загрузить лист с именем "Data" в переменную sheet

def getvalue(x):
    return x.value
years = list(map(getvalue, sheet['A'][1:])) # Преобразовать содержимое колонки A в список в переменной years, с помощью list и map
temp = list(map(getvalue, sheet['C'][1:])) # Преобразовать содержимое колонки С в список в переменной temp,с помощью list и map
sun = list(map(getvalue, sheet['D'][1:])) # Преобразовать содержимое колонки D в список в переменной sun, с помощью list и map
pyplot.plot(years, temp, label="Температура") # Построить график по точкам X=years, Y=temp
pyplot.plot(years, sun, label="Активность солнца") # Построить график по точкам X=years, Y=sun
pyplot.show() # Показать график

